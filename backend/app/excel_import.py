"""뱅크샐러드 가계부 엑셀("가계부 내역" 시트) 파싱.

행 변환 규칙:
- 이체 타입은 자동 반영하지 않고 "검토 대상"으로 분리 반환한다 — 행마다 한쪽
  다리(결제수단 계정)만 기록되고 상대 계정은 자유 텍스트뿐이라, 업로드 확정
  시점에 사용자가 수입/지출/이체(상대 계정 지정)/건너뛰기를 결정한다.
  단, 내계좌이체끼리 같은 날짜·같은 금액·반대 부호·다른 결제수단이면 자동
  페어링해 한 건의 이체(출금→입금)로 제안한다.
- 지출+양수(환불)는 수입으로 반영하되 카테고리는 '환불 > 미분류', 원래 분류는 memo에 보존.
- 금액 0원, KRW 외 통화, 타입과 부호가 모순인 행은 스킵하고 사유를 남긴다.
"""

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import load_workbook

SHEET_NAME = "가계부 내역"
REQUIRED_COLUMNS = ("날짜", "타입", "대분류", "소분류", "내용", "금액", "화폐", "결제수단")
UNCLASSIFIED = "미분류"
REFUND_MAJOR = "환불"
# 엑셀 날짜 시리얼 기준일 (1900 윤년 버그 보정 포함)
EXCEL_EPOCH = date(1899, 12, 30)
MEMO_MAX = 255


class ExcelFormatError(ValueError):
    """시트/헤더가 기대 형식이 아닐 때 — 라우터에서 422로 변환한다."""


@dataclass
class ParsedRow:
    row: int  # 엑셀 행 번호 (헤더 포함 1부터)
    date: date
    kind: str  # income | expense
    major: str
    minor: str
    amount: int  # 항상 양수
    account_name: str
    memo: str | None


@dataclass
class SkippedRow:
    row: int
    reason: str


@dataclass
class ParsedValuation:
    """뱅샐현황 "재무현황" 자산 표에서 추출한 시세형 자산 평가액 한 줄."""

    product_name: str  # 엑셀 상품명 — 자산 계정명과 매칭한다
    account_type: str  # real_estate | stock
    value: int  # KRW 정수(원), 0 이상


@dataclass
class ReviewRow:
    """이체 타입 행 — 자동 반영하지 않고 사용자 결정을 기다리는 검토 대상.

    amount는 부호를 보존한다 (음수=결제수단 계정에서 출금, 양수=입금).
    pair_row는 내계좌이체 자동 페어링 결과 — 상대 다리의 엑셀 행 번호.
    """

    row: int
    date: date
    major: str
    minor: str
    description: str | None
    amount: int
    account_name: str
    pair_row: int | None = None


def _to_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return EXCEL_EPOCH + timedelta(days=int(value))
    return None


def _clean(value) -> str:
    return str(value).strip() if value is not None else ""


def parse_ledger(
    content: bytes, month: str
) -> tuple[list[ParsedRow], list[ReviewRow], list[SkippedRow], int]:
    """지정 월(YYYY-MM)의 행을 파싱한다.

    returns: (유효 행 목록, 이체 검토 행 목록, 스킵 행 목록, 해당 월 전체 행 수)
    raises: ExcelFormatError — 시트 부재, 필수 컬럼 누락 등 파일 형식 문제
    """
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl은 손상/비xlsx 파일에 다양한 예외를 던진다
        raise ExcelFormatError("엑셀 파일을 열 수 없습니다 (.xlsx 형식인지 확인해주세요)") from exc

    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ExcelFormatError(f"'{SHEET_NAME}' 시트를 찾을 수 없습니다")
        sheet = workbook[SHEET_NAME]

        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ExcelFormatError(f"'{SHEET_NAME}' 시트가 비어 있습니다")
        col = {_clean(name): idx for idx, name in enumerate(header) if name is not None}
        missing = [name for name in REQUIRED_COLUMNS if name not in col]
        if missing:
            raise ExcelFormatError(f"필수 컬럼이 없습니다: {', '.join(missing)}")
        memo_idx = col.get("메모")

        parsed: list[ParsedRow] = []
        review: list[ReviewRow] = []
        skipped: list[SkippedRow] = []
        month_rows = 0

        for row_no, values in enumerate(rows, start=2):
            def cell(name: str):
                idx = col[name]
                return values[idx] if idx < len(values) else None

            tx_date = _to_date(cell("날짜"))
            if tx_date is None:
                continue  # 날짜 없는 행(빈 행 등)은 월 판정 자체가 불가 — 조용히 무시
            if tx_date.strftime("%Y-%m") != month:
                continue
            month_rows += 1

            def skip(reason: str):
                skipped.append(SkippedRow(row=row_no, reason=reason))

            currency = _clean(cell("화폐"))
            if currency and currency != "KRW":
                skip(f"KRW 외 통화({currency})는 지원하지 않습니다")
                continue

            tx_type = _clean(cell("타입"))
            if tx_type not in ("지출", "수입", "이체"):
                skip(f"알 수 없는 타입입니다: {tx_type or '(빈 값)'}")
                continue

            raw_amount = cell("금액")
            if not isinstance(raw_amount, (int, float)) or int(raw_amount) == 0:
                skip("금액이 0원이거나 숫자가 아닙니다")
                continue
            amount = int(raw_amount)

            major = _clean(cell("대분류")) or UNCLASSIFIED
            minor = _clean(cell("소분류")) or UNCLASSIFIED
            memo = _clean(cell("내용"))
            excel_memo = _clean(values[memo_idx] if memo_idx is not None and memo_idx < len(values) else None)
            if excel_memo:
                memo = f"{memo} — {excel_memo}" if memo else excel_memo

            if tx_type == "이체":
                # 자동 반영하지 않고 검토 대상으로 — 부호(입출금 방향)를 보존한다
                review.append(
                    ReviewRow(
                        row=row_no,
                        date=tx_date,
                        major=major,
                        minor=minor,
                        description=memo[:MEMO_MAX] if memo else None,
                        amount=amount,
                        account_name=_clean(cell("결제수단")) or "미지정",
                    )
                )
                continue

            if tx_type == "수입" and amount < 0:
                skip("수입인데 금액이 음수입니다")
                continue

            if tx_type == "지출" and amount > 0:
                # 환불 — 수입으로 반영하고 원래 분류는 memo에 보존
                origin = major if minor == UNCLASSIFIED else f"{major} > {minor}"
                memo = f"{memo} [환불: {origin}]" if memo else f"[환불: {origin}]"
                kind, major, minor = "income", REFUND_MAJOR, UNCLASSIFIED
            else:
                kind = "income" if tx_type == "수입" else "expense"

            parsed.append(
                ParsedRow(
                    row=row_no,
                    date=tx_date,
                    kind=kind,
                    major=major,
                    minor=minor,
                    amount=abs(amount),
                    account_name=_clean(cell("결제수단")) or "미지정",
                    memo=memo[:MEMO_MAX] if memo else None,
                )
            )

        _pair_own_transfers(review)
        return parsed, review, skipped, month_rows
    finally:
        workbook.close()


OWN_TRANSFER_MAJOR = "내계좌이체"


def _pair_own_transfers(review: list[ReviewRow]) -> None:
    """내계좌이체 행의 출금(-)·입금(+) 다리를 1:1 페어링한다.

    조건: 같은 날짜, 같은 절대 금액, 반대 부호, 다른 결제수단.
    페어된 두 행은 pair_row로 서로를 가리키며, 확정 시 한 건의 이체가 된다.
    """
    own = [r for r in review if r.major == OWN_TRANSFER_MAJOR]
    incoming = [r for r in own if r.amount > 0]
    used: set[int] = set()
    for out in (r for r in own if r.amount < 0):
        for i, inc in enumerate(incoming):
            if i in used:
                continue
            if (
                inc.date == out.date
                and inc.amount == -out.amount
                and inc.account_name != out.account_name
            ):
                out.pair_row, inc.pair_row = inc.row, out.row
                used.add(i)
                break


def guess_account_type(name: str) -> str:
    """결제수단명으로 자산 계정 type을 추정한다 (보수적 휴리스틱)."""
    lowered = name.lower()
    if "카드" in name or "check" in lowered or "체크" in name:
        return "card"
    if "통장" in name or "은행" in name or "뱅크" in name:
        return "bank"
    if "현금" in name:
        return "cash"
    return "other"


# 평가액을 읽는 첫 시트와 자산 분류→계정 유형 매핑
VALUATION_SHEET = "뱅샐현황"
VALUATION_ITEM_TYPES = {"부동산": "real_estate", "투자성 자산": "stock"}


def parse_valuations(content: bytes) -> list[ParsedValuation]:
    """뱅샐현황 시트의 "재무현황" 자산 표에서 부동산·주식 평가액을 추출한다.

    항목(자산 분류) 셀은 그룹의 첫 행에만 있고 이후 행은 엑셀 병합으로 비어 있으므로
    마지막 분류를 carry-forward 한다. 부동산→real_estate, 투자성 자산→stock으로 매핑하고,
    '총자산'/'순자산' 등 집계 행이나 다음 섹션 헤더에서 표를 종료한다. 행 위치는 파일마다
    다를 수 있어 헤더(항목/상품명/금액)를 탐지해 컬럼을 잡는다.

    시트가 없거나 자산 표를 찾지 못하면 빈 목록을 반환한다 — 평가 반영은 선택 사항이므로
    가져오기 자체를 막지 않는다. 0원 행도 그대로 반환하며, 신규 생성/갱신 정책은 호출 측이 정한다.
    """
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:  # 손상/비xlsx 등 — 평가는 선택 사항이므로 조용히 건너뛴다
        return []

    try:
        if VALUATION_SHEET not in workbook.sheetnames:
            return []
        sheet = workbook[VALUATION_SHEET]
        rows = sheet.iter_rows(values_only=True)

        # 자산 표 헤더(항목/상품명/금액)를 찾아 컬럼 인덱스를 잡는다 (부채 측 동명 컬럼은 무시)
        item_col = product_col = amount_col = None
        for values in rows:
            cleaned = [_clean(v) for v in values]
            if "항목" in cleaned and "상품명" in cleaned and "금액" in cleaned:
                try:
                    item_col = cleaned.index("항목")
                    product_col = cleaned.index("상품명", item_col + 1)
                    amount_col = cleaned.index("금액", product_col + 1)
                except ValueError:
                    # 항목→상품명→금액 순서가 아니면 자산 표 헤더가 아님 — 다음 행 탐색
                    item_col = None
                    continue
                break
        if item_col is None:
            return []

        result: list[ParsedValuation] = []
        current_item = ""
        for values in rows:  # 헤더 다음 행부터 (제너레이터가 이미 헤더를 소비함)
            raw_item = _clean(values[item_col]) if item_col < len(values) else ""
            if raw_item in ("총자산", "순자산"):
                break  # 자산 표 끝
            if raw_item[:1].isdigit() and "." in raw_item[:3]:
                break  # 다음 섹션 헤더(예: '4.보험현황')
            if raw_item:
                current_item = raw_item
            account_type = VALUATION_ITEM_TYPES.get(current_item)
            if account_type is None:
                continue
            product = _clean(values[product_col]) if product_col < len(values) else ""
            amount = values[amount_col] if amount_col < len(values) else None
            if not product or not isinstance(amount, (int, float)):
                continue
            value = int(amount)
            if value < 0:
                continue
            result.append(
                ParsedValuation(product_name=product, account_type=account_type, value=value)
            )
        return result
    finally:
        workbook.close()
